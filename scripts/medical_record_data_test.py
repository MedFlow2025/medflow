# Copyright (c) 2025,  IEIT SYSTEMS Co.,Ltd.  All rights reserved

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#     http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import json
import requests
from openpyxl import Workbook
from openpyxl.styles import Font

# 读取JSON文件
with open('medical_record_data.json', 'r', encoding='utf-8') as json_file:
    data_list = json.load(json_file)
    
with open('../quality/quality.json', 'r', encoding='utf-8') as json_file:
    quality_info = json.load(json_file)
    
    quality_list = quality_info['check_quality']
    print(f"***lmx*** quality_list {quality_info} , quality_list {quality_list}")
    
url = 'http://ip:port/quality_inspect'
headers = {
    'accept': 'application/json',
    'Content-Type': 'application/json'
}

# 创建Excel工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = "MedicalData"

headers_excel = ['response_status_code', 'response_text', 'content', 'field', 'item', 'standard', 'check_quality', 'auto_modify_type', 'auto_modify_info', 'other_info']

input_base_medical_head = ['chief_complaint', 'history_of_present_illness', 'past_medical_history', 'personal_history',  'allergy_history', 'physical_examination', 'auxiliary_examination']
input_physical_examination_head = ['temperature', 'pulse', 'blood_pressure', 'respiration']

for col_num, header in enumerate(headers_excel, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

row_num = 2
for data in data_list:
    try:
        json_data = json.dumps(data, ensure_ascii=True)
        print(f"***lmx*** json_data {json_data} , row_num {row_num}, data {data}")
        response = requests.post(url, headers=headers, data=json_data, timeout=50)
        
        ws.cell(row=row_num, column=1).value = response.status_code
        ws.cell(row=row_num, column=2).value = response.text
        
        
        response_data = json.loads(response.text)
        print(f"***lmx*** response_data {json.dumps(response_data, ensure_ascii=False)}")  
        
        row_input_info = row_num
        response_base_medical = response_data['output']['basic_medical_record']
        for base_info in input_base_medical_head:
            row_input_info = row_input_info + 1
            ws.cell(row=row_input_info, column=1).value = base_info
            ws.cell(row=row_input_info, column=2).value = str(response_base_medical[base_info])
        
        response_physical_examination = response_data['output']['basic_medical_record']['physical_examination']
        for base_info in input_physical_examination_head:
            row_input_info = row_input_info + 1
            ws.cell(row=row_input_info, column=1).value = base_info
            ws.cell(row=row_input_info, column=2).value = response_physical_examination[base_info]


        response_quality_list = response_data['output']['control_quality']
        response_col = 1
        
        for response_info in response_quality_list:
            row_num = row_num + 1
            ws.cell(row=row_num, column=3).value = response_info['content']
            ws.cell(row=row_num, column=4).value = response_info['field']
            ws.cell(row=row_num, column=5).value = response_info['item']
            ws.cell(row=row_num, column=6).value = response_info['standard']
            ws.cell(row=row_num, column=7).value = response_info['check_quality']
            ws.cell(row=row_num, column=8).value = response_info['auto_modify_type']
            ws.cell(row=row_num, column=9).value = response_info['auto_modify_info']
        
        row_num = row_num + 1
    except Exception as e:
        print(f"发生异常: {e}")

wb.save('output.xlsx')
'''
# 设置表头
headers_excel = ['chief_complaint', 'history_of_present_illness', 'past_medical_history', 'personal_history', 'allergy_history',
                'temperature', 'pulse', 'blood_pressure', 'respiration', 'auxiliary_examination','response_status_code','response_text']
for col_num, header in enumerate(headers_excel, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)


sheet2 = wb.create_sheet("result_check")
result_check_headers_excel = []
for cq in quality_list:
    result_check_headers_excel.append(cq['content'])
    user_content = "check_" + cq['content']
    result_check_headers_excel.append(user_content)
    result_check_headers_excel.append("result")
    
for col_num, header in enumerate(result_check_headers_excel, 1):
    cell = sheet2.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)


row_num = 2
for data in data_list:
    try:
        json_data = json.dumps(data)
        print(f"***lmx*** json_data {json_data} , row_num {row_num}, data {data}")
        
        response = requests.post(url, headers=headers, data=json_data, timeout=50)
        print(f"***lmx*** response {response}  {response.text}")
        basic_medical_record = data['input']['basic_medical_record']
        ws.cell(row=row_num, column=1).value = basic_medical_record['chief_complaint']
        ws.cell(row=row_num, column=2).value = basic_medical_record['history_of_present_illness']
        ws.cell(row=row_num, column=3).value = basic_medical_record['past_medical_history']
        ws.cell(row=row_num, column=4).value = basic_medical_record['personal_history']
        ws.cell(row=row_num, column=5).value = basic_medical_record['allergy_history']
        ws.cell(row=row_num, column=6).value = basic_medical_record['physical_examination']['temperature']
        ws.cell(row=row_num, column=7).value = basic_medical_record['physical_examination']['pulse']
        ws.cell(row=row_num, column=8).value = basic_medical_record['physical_examination']['blood_pressure']
        ws.cell(row=row_num, column=9).value = basic_medical_record['physical_examination']['respiration']
        ws.cell(row=row_num, column=10).value = basic_medical_record['auxiliary_examination']
        ws.cell(row=row_num, column=11).value = response.status_code
        ws.cell(row=row_num, column=12).value = response.text
        
        
        response_data = json.loads(response.text)
        # print(f"***lmx*** response_data {response_data}")
        response_quality_list = response_data['output']['control_quality']
        response_col = 1
        for response_info in response_quality_list:
            result = response_info['check_quality']
            if response_info['auto_modify_type'] and len(response_info['auto_modify_info']) > 0:
                result = result + "_" + response_info['auto_modify_info']
            
            response_col_temp = 1    
            for index, element in enumerate(result_check_headers_excel):
                if element == response_info['content']:
                    response_col_temp = index
                    break
            response_col_temp = response_col_temp + 1
            sheet2.cell(row=row_num, column=response_col_temp).value = result
            
        row_num += 1

    except Exception as e:
        print(f"发生异常: {e}")
#'''

