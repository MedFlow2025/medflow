# Copyright (c) 2025,  IEIT SYSTEMS Co.,Ltd.  All rights reserved

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#     http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import copy
import json
import math
import requests
from openpyxl import Workbook
from openpyxl.styles import Font

# 读取JSON文件
with open('medical_record_data.json', 'r', encoding='utf-8') as json_file:
    data_list = json.load(json_file)
    
with open('../quality/quality.json', 'r', encoding='utf-8') as json_file:
    quality_info = json.load(json_file)
    
    quality_list = quality_info['check_quality']
    print(f"***lmx*** quality_list {quality_info} , quality_list {quality_list}")
    
url = 'http://ip:port/quality_inspect'
headers = {
    'accept': 'application/json',
    'Content-Type': 'application/json'
}
url_modify = 'http://ip:port/quality_modify'

# 创建Excel工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = "MedicalData"

headers_excel = ['response_status_code', 'response_text', 'content', 'field', 'item', 'standard', 'check_quality', 'auto_modify_type', 'auto_modify_info', 'other_info']

input_base_medical_head = ['chief_complaint', 'history_of_present_illness', 'past_medical_history', 'personal_history',  'allergy_history', 'physical_examination', 'auxiliary_examination']
input_physical_examination_head = ['temperature', 'pulse', 'blood_pressure', 'respiration']

for col_num, header in enumerate(headers_excel, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

row_num = 2
for data in data_list:
    try:
        base_row_num = row_num
        max_row_num = row_num
        json_data = json.dumps(data, ensure_ascii=True)
        print(f"***lmx*** json_data {json_data} , row_num {row_num}, data {data}")
        response = requests.post(url, headers=headers, data=json_data, timeout=50)
        
        ws.cell(row=row_num, column=1).value = response.status_code
        ws.cell(row=row_num, column=2).value = response.text
        
        
        response_data = json.loads(response.text)
        print(f"***lmx*** response_data {json.dumps(response_data, ensure_ascii=False)}")  
        
        # 写待测试输入内容信息
        row_input_info = row_num
        response_base_medical = response_data['output']['basic_medical_record']
        for base_info in input_base_medical_head:
            row_input_info = row_input_info + 1
            ws.cell(row=row_input_info, column=1).value = base_info
            ws.cell(row=row_input_info, column=2).value = str(response_base_medical[base_info])
        
        response_physical_examination = response_data['output']['basic_medical_record']['physical_examination']
        for base_info in input_physical_examination_head:
            row_input_info = row_input_info + 1
            ws.cell(row=row_input_info, column=1).value = base_info
            ws.cell(row=row_input_info, column=2).value = response_physical_examination[base_info]
        
        max_row_num = max(max_row_num, row_input_info)


        # 写质检回复内容
        response_quality_list = response_data['output']['control_quality']
        response_col = 1
        deep_copied_data = copy.deepcopy(data)
        response_base_medical_modify = copy.deepcopy(response_data['output']['basic_medical_record'])
        for response_info in response_quality_list:
            row_num = row_num + 1
            ws.cell(row=row_num, column=3).value = response_info['content']
            ws.cell(row=row_num, column=4).value = response_info['field']
            ws.cell(row=row_num, column=5).value = response_info['item']
            ws.cell(row=row_num, column=6).value = response_info['standard']
            ws.cell(row=row_num, column=7).value = response_info['check_quality']
            ws.cell(row=row_num, column=8).value = response_info['auto_modify_type']
            ws.cell(row=row_num, column=9).value = response_info['auto_modify_info']
            
            print(f"***lmx*** response_info {response_info}")
            
            # modify 接口部分调用和获取数据 检查是否存在单位不全
            if response_info['auto_modify_type'] == True:
                deep_copied_data = copy.deepcopy(deep_copied_data)
                deep_copied_data['input']['basic_medical_record'] = response_base_medical_modify
                deep_copied_data['input']['control_quality'] = []
                deep_copied_data['input']['control_quality'].append(response_info)
                deep_copied_data['input']['confirm_auto_modify'] = True
                
                deep_copied_data['chat'] = {}
                deep_copied_data['chat']['historical_conversations'] = []
                
                json_copied_data = json.dumps(deep_copied_data, ensure_ascii=False)
                response_auto_modify = requests.post(url_modify, headers=headers, data=json_copied_data, timeout=50)
                  
                response_data_modify = json.loads(response_auto_modify.text)
                response_base_medical_modify = copy.deepcopy(response_data_modify['output']['basic_medical_record'])
                print(f"***lmx*** json_copied_data {json_copied_data} \n response_auto_modify {response_auto_modify.text} \n response_base_medical_modify {response_base_medical_modify}")
                
            # modify 接口，处理为空的部分错误
            if "是否为空" in response_info['content']:
                deep_copied_data = copy.deepcopy(deep_copied_data)
                deep_copied_data['input']['basic_medical_record'] = response_base_medical_modify
                deep_copied_data['input']['control_quality'] = []
                deep_copied_data['input']['control_quality'].append(response_info)
                deep_copied_data['input']['confirm_auto_modify'] = False
                
                deep_copied_data['chat'] = {}
                deep_copied_data['chat']['historical_conversations'] = []
                
                
                json_copied_data = json.dumps(deep_copied_data, ensure_ascii=False)
                response_auto_modify = requests.post(url_modify, headers=headers, data=json_copied_data, timeout=50)
                print(f"***lmx*** kong response_auto_modify {response_auto_modify.text}  \n deep_copied_data {deep_copied_data}")  
                response_data_modify = json.loads(response_auto_modify.text)
                deep_copied_data['input'] = response_data_modify['output']
                deep_copied_data['chat'] = response_data_modify['chat']
                # temp_content = response_info['field'] + " "  + " 的补充内容如下： 这里是新加的数据。" 
                temp_content = "补充内容如下： 新添加数据。" 
                temp_dict = {"role": "user", "content": temp_content}
                deep_copied_data['chat']['historical_conversations'].append(temp_dict)

                json_copied_data = json.dumps(deep_copied_data, ensure_ascii=False)
                response_auto_modify = requests.post(url_modify, headers=headers, data=json_copied_data, timeout=50)
                
                response_data_modify = json.loads(response_auto_modify.text)
                response_base_medical_modify = copy.deepcopy(response_data_modify['output']['basic_medical_record'])
                print(f"***lmx*** kong 2  json_copied_data {json_copied_data} \n response_auto_modify {response_auto_modify.text}  \n response_base_medical_modify {response_base_medical_modify}")  

            # 数值部分修改
            if "数值" in response_info['content']:
                deep_copied_data = copy.deepcopy(deep_copied_data)
                deep_copied_data['input']['basic_medical_record'] = response_base_medical_modify
                deep_copied_data['input']['control_quality'] = []
                deep_copied_data['input']['control_quality'].append(response_info)
                deep_copied_data['input']['confirm_auto_modify'] = False
                
                deep_copied_data['chat'] = {}
                deep_copied_data['chat']['historical_conversations'] = []
                
                json_copied_data = json.dumps(deep_copied_data, ensure_ascii=False)
                response_auto_modify = requests.post(url_modify, headers=headers, data=json_copied_data, timeout=50)
                print(f"***lmx*** value response_auto_modify {response_auto_modify.text}")  
                response_data_modify = json.loads(response_auto_modify.text)
                deep_copied_data['input'] = response_data_modify['output']
                deep_copied_data['chat'] = response_data_modify['chat']
                
                temp_value = ""
                if response_info['item'] == "体温":
                    temp_value = "36.7℃"
                if response_info['item'] == "脉搏":
                    temp_value = "70"
                if response_info['item'] == "血压":
                    temp_value = "120/80mmHg"
                if response_info['item'] == "呼吸":
                    temp_value = "19 次/分"                

                temp_content = response_info['field'] + " "  + " 修改后数值如下："  + temp_value
                temp_dict = {"role": "user", "content": temp_content}
                deep_copied_data['chat']['historical_conversations'].append(temp_dict)

                json_copied_data = json.dumps(deep_copied_data, ensure_ascii=False)
                response_auto_modify = requests.post(url_modify, headers=headers, data=json_copied_data, timeout=50)
                
                response_data_modify = json.loads(response_auto_modify.text)
                response_base_medical_modify = copy.deepcopy(response_data_modify['output']['basic_medical_record'])
                print(f"***lmx*** value 2 json_copied_data {json_copied_data} \n response_auto_modify {response_auto_modify.text} \n response_base_medical_modify {response_base_medical_modify}")  
        
        max_row_num = max(max_row_num, row_num)        

        # 重新写一下病例内容
        # 写待测试输入内容信息
        row_input_info = base_row_num
        for base_info in input_base_medical_head:
            row_input_info = row_input_info + 1
            ws.cell(row=row_input_info, column=11).value = str(response_base_medical_modify[base_info])
        
        response_physical_examination = response_base_medical_modify['physical_examination']
        for base_info in input_physical_examination_head:
            row_input_info = row_input_info + 1
            ws.cell(row=row_input_info, column=11).value = response_physical_examination[base_info]
        
        max_row_num = max(max_row_num, row_input_info)
        row_num = max_row_num + 1
    except Exception as e:
        print(f"发生异常: {e}")

wb.save('quality_modiyf_output.xlsx')

