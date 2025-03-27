# Copyright (c) 2025,  IEIT SYSTEMS Co.,Ltd.  All rights reserved

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#     http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import json
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
import argparse

# 使用示例
# python medical_record_data_test.py  --medical-record  medical_record_data_debug.json    --quality-name   quality_base  
# --output-file quality_base_debug_out.xlsx

def args_parser():
    parser = argparse.ArgumentParser(description='Quality test with Customizable Parameters')
    parser.add_argument("--host", type=str, default="127.0.0.1", help="The inference server ip")
    parser.add_argument("--port", type=int, default="8013")
    parser.add_argument('--medical-record', type=str, default="medical_record_data.json", help='medical_record file name')
    parser.add_argument('--quality-file', type=str, default="../database/quality/quality.json", help='select quality file')
    parser.add_argument('--quality-name', type=str, default=None, help='select quality name')
    parser.add_argument('--output-file', type=str, default="output.xlsx", help='output quality check file name')
    args = parser.parse_args()
    return args

args = args_parser()


# 读取JSON文件
with open(args.medical_record, 'r', encoding='utf-8') as json_file:
    data_list = json.load(json_file)
    
with open(args.quality_file, 'r', encoding='utf-8') as json_file:
    quality_info = json.load(json_file)
    
    quality_list = quality_info['check_quality']
    print(f"***lmx*** quality_list {quality_info} , quality_list {quality_list}")
    
url = 'http://ip:port/quality_inspect'
headers = {
    'accept': 'application/json',
    'Content-Type': 'application/json'
}

# 创建Excel工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = "MedicalData"

headers_excel = ['response_status_code', 'response_text', 'content', 'field', 'item', 'standard', 'check_quality', 'auto_modify_type', 'auto_modify_info', 'other_info']

input_base_medical_head = ['chief_complaint', 'history_of_present_illness', 'past_medical_history', 'personal_history',  'allergy_history', 'physical_examination', 'auxiliary_examination']
input_physical_examination_head = ['temperature', 'pulse', 'blood_pressure', 'respiration']

for col_num, header in enumerate(headers_excel, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

row_num = 2
for data in data_list:
    try:
        if args.quality_name is not None:
            data["input"]["control_quality_config_name"] = args.quality_name        
        json_data = json.dumps(data, ensure_ascii=True)

        print(f"***lmx*** json_data {json_data} , row_num {row_num}, data {data}")
        response = requests.post(url, headers=headers, data=json_data, timeout=50)
        
        ws.cell(row=row_num, column=1).value = response.status_code
        ws.cell(row=row_num, column=2).value = response.text
        
        
        response_data = json.loads(response.text)
        print(f"***lmx*** response_data {json.dumps(response_data, ensure_ascii=False)}")  
        
        row_input_info = row_num
        response_base_medical = response_data['output']['basic_medical_record']
        for base_info in input_base_medical_head:
            row_input_info = row_input_info + 1
            ws.cell(row=row_input_info, column=1).value = base_info
            ws.cell(row=row_input_info, column=2).value = str(response_base_medical[base_info])
        
        response_physical_examination = response_data['output']['basic_medical_record']['physical_examination']
        for base_info in input_physical_examination_head:
            row_input_info = row_input_info + 1
            ws.cell(row=row_input_info, column=1).value = base_info
            ws.cell(row=row_input_info, column=2).value = response_physical_examination[base_info]


        response_quality_list = response_data['output']['control_quality']
        response_col = 1
        
        for response_info in response_quality_list:
            row_num = row_num + 1
            ws.cell(row=row_num, column=3).value = response_info['content']
            ws.cell(row=row_num, column=4).value = response_info['field']
            ws.cell(row=row_num, column=5).value = response_info['item']
            ws.cell(row=row_num, column=6).value = response_info['standard']
            ws.cell(row=row_num, column=7).value = response_info['check_quality']
            ws.cell(row=row_num, column=8).value = response_info['auto_modify_type']
            ws.cell(row=row_num, column=9).value = response_info['auto_modify_info']
        
        max_row_num = max(row_input_info, row_num)
        row_num = max_row_num + 1
    except Exception as e:
        print(f"发生异常: {e}")

wb.save(args.output_file)

